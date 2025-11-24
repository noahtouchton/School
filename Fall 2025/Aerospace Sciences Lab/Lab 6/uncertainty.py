# uncertainty.py
import os
import pandas as pd
import numpy as np
from rss import Data, RSS, linear_monte_carlo
import sympy as sp

# ------------------------------------------------------------
# Equations used with RSS
# ------------------------------------------------------------
equations = [
    "ρ = P / (R * T)",            # 0: density from ambient P,T
    "A = 3.14159 * d**2 / 4",     # 1: orifice area
    "a = sqrt(gamma * R * T0)",   # 2: speed of sound from T0 (3-state)
    "tchar = V / (A * a)",        # 3: characteristic time
    "Pabs = Pg + P",          # 4: absolute pressure from gauge + ambient
]

# ------------------------------------------------------------
# Lab ambient & geometry (edit if needed)
# ------------------------------------------------------------
d_values = np.array([0.0017018, 0.0023876, 0.003175])        # small, medium, large [m]
d_uncertainties = np.array([2.54e-05, 2.54e-05, 2.54e-05])   # ±0.001 in in meters

T_amb = Data("Ambient Temperature", "T", 297.05, 0.4)        # K
P_amb = Data("Ambient Pressure",    "P", 101500.0, 400.0)    # Pa
R     = Data("Ideal Gas Constant",  "R", 287.0, 0.0)         # J/kg/K
gamma = Data("Specific Heat Ratio", "gamma", 1.4, 0.0)

d = Data("Orifice Diameter", "d", d_values, d_uncertainties)
A = RSS("Orifice Area", equations[1], [d])                   # 3-state: small, medium, large

V = Data("Tank Volume", "V", 0.125, 0.0)                     # m^3

# Derived ambient density (scalar)
ρ = RSS("Density", equations[0], [P_amb, R, T_amb])


# ------------------------------------------------------------
# Data loading from Lab6Data.xlsx
# ------------------------------------------------------------
def load_lab6_data():
    """
    Load large, medium, and small orifice runs from Lab6Data.xlsx.

    Columns in each block:
      A: time [s]
      B: gauge pressure [kPa]
      C: absolute pressure [kPa]
      D: temperature [°C]
    """
    here = os.path.dirname(os.path.abspath(__file__))
    excel_path = os.path.join(here, "Lab6Data.xlsx")

    from openpyxl import load_workbook
    wb = load_workbook(excel_path, data_only=True)
    ws = wb[wb.sheetnames[0]]  # first sheet

    def read_range(range_str):
        start_cell, end_cell = range_str.split(":")
        start_row = int(start_cell[1:])
        end_row   = int(end_cell[1:])

        data_rows = []
        for r in range(start_row, end_row + 1):
            row_vals = []
            for c in ["A", "B", "C", "D"]:
                cell = f"{c}{r}"
                row_vals.append(ws[cell].value)
            data_rows.append(row_vals)

        df = pd.DataFrame(data_rows, columns=["A", "B", "C", "D"])
        return df

    # Ranges (as you specified)
    df_large  = read_range("A143:D206")   # large orifice
    df_medium = read_range("A302:D424")   # medium orifice
    df_small  = read_range("A734:D918")   # small orifice

    def convert_block(df):
        t    = np.array(df["A"],    dtype=float)  # time [s]
        Pg_k = np.array(df["B"],    dtype=float)  # gauge [kPa]
        Pabs_k = np.array(df["C"],  dtype=float)  # absolute [kPa]
        Tc   = np.array(df["D"],    dtype=float)  # temp [°C]

        Pg   = Pg_k   * 1000.0      # Pa
        Pabs = Pabs_k * 1000.0      # Pa
        T    = Tc + 273.15          # K

        return t, Pg, Pabs, T

    t_large, Pg_large, Pabs_large, T_large = convert_block(df_large)
    t_med,   Pg_med,   Pabs_med,   T_med   = convert_block(df_medium)
    t_small, Pg_small, Pabs_small, T_small = convert_block(df_small)

    # Combined stacks if needed
    t_all    = np.concatenate([t_large, t_med, t_small])
    Pg_all   = np.concatenate([Pg_large, Pg_med, Pg_small])
    Pabs_all = np.concatenate([Pabs_large, Pabs_med, Pabs_small])
    T_all    = np.concatenate([T_large, T_med, T_small])

    return {
        "t_large": t_large,
        "Pg_large": Pg_large,
        "Pabs_large": Pabs_large,
        "T_large": T_large,

        "t_med": t_med,
        "Pg_med": Pg_med,
        "Pabs_med": Pabs_med,
        "T_med": T_med,

        "t_small": t_small,
        "Pg_small": Pg_small,
        "Pabs_small": Pabs_small,
        "T_small": T_small,

        "t_all": t_all,
        "Pg_all": Pg_all,
        "Pabs_all": Pabs_all,
        "T_all": T_all,
    }


data = load_lab6_data()

# ------------------------------------------------------------
# Uncertainties from lab
# ------------------------------------------------------------
P_spec_unc = 1724.0  # Pa spec for pressure transducer (±0.25% FS of 100 psia)
T_unc = 0.5          # K
t_unc = 0.0          # s

# ------------------------------------------------------------
# Calibration curve via linear_monte_carlo
# x = sensor reading (e.g. V), y = pressure [Pa]
# ------------------------------------------------------------
x_vals = np.array([1.438, 1.8125, 1.88, 2.5938, 2.9062, 3.060, 3.313])
y_vals_kPa = np.array([100, 151, 200, 252, 302, 353, 405])

x_cal = Data("Calibration x", "x",
             x_vals,
             np.full_like(x_vals, 0.0625, dtype=float))   # sensor resolution

y_cal = Data("Calibration y", "y",
             y_vals_kPa * 1000.0,                         # kPa -> Pa
             np.full_like(y_vals_kPa, P_spec_unc, dtype=float))

slope, intercept = linear_monte_carlo(x_cal, y_cal, N=100_000)
# slope.var.name = "m", intercept.var.name = "b"

# ------------------------------------------------------------
# Recover effective sensor readings (x_eff) from calibrated Pg
# and then recompute Pg with RSS to get calibration-based σ_P
# ------------------------------------------------------------
m_val = slope.value
b_val = intercept.value
x_unc = 0.0625  # same as calibration x uncertainty

# --- LARGE ORIFICE ---
PgL_nom = np.asarray(data["Pg_large"], dtype=float)   # Pa
xL_eff  = (PgL_nom - b_val) / m_val                   # invert P = m x + b

X_large = Data("Recovered Sensor Large", "x",
               xL_eff,
               np.full_like(xL_eff, x_unc, dtype=float))

calib_eq = "Pg = m * x + b"
Pg_large = RSS("Gauge Pressure Large (calibrated)",
               calib_eq,
               [slope, intercept, X_large])
Pabs_large = RSS("Absolute Pressure Large",
                 equations[4],
                    [Pg_large, P_amb])

# --- MEDIUM ORIFICE ---
PgM_nom = np.asarray(data["Pg_med"], dtype=float)
xM_eff  = (PgM_nom - b_val) / m_val

X_med = Data("Recovered Sensor Medium", "x",
             xM_eff,
             np.full_like(xM_eff, x_unc, dtype=float))

Pg_med = RSS("Gauge Pressure Medium (calibrated)",
             calib_eq,
             [slope, intercept, X_med])
Pabs_med = RSS("Absolute Pressure Medium",
               equations[4],
                [Pg_med, P_amb])

# --- SMALL ORIFICE ---
PgS_nom = np.asarray(data["Pg_small"], dtype=float)
xS_eff  = (PgS_nom - b_val) / m_val

X_small = Data("Recovered Sensor Small", "x",
               xS_eff,
               np.full_like(xS_eff, x_unc, dtype=float))

Pg_small = RSS("Gauge Pressure Small (calibrated)",
               calib_eq,
               [slope, intercept, X_small])
Pabs_small = RSS("Absolute Pressure Small",
                    equations[4],
                    [Pg_small, P_amb])

# ------------------------------------------------------------
# Time and Temperature Data objects for each orifice
# ------------------------------------------------------------
t_large = Data("Time Large",  "t_large",
               data["t_large"],
               np.full_like(data["t_large"], t_unc, dtype=float))

t_med   = Data("Time Medium", "t_med",
               data["t_med"],
               np.full_like(data["t_med"], t_unc, dtype=float))

t_small = Data("Time Small",  "t_small",
               data["t_small"],
               np.full_like(data["t_small"], t_unc, dtype=float))

T_large = Data("Temperature Large", "T_large",
               data["T_large"],
               np.full_like(data["T_large"], T_unc, dtype=float))

T_med   = Data("Temperature Medium", "T_med",
               data["T_med"],
               np.full_like(data["T_med"], T_unc, dtype=float))

T_small = Data("Temperature Small", "T_small",
               data["T_small"],
               np.full_like(data["T_small"], T_unc, dtype=float))

# ------------------------------------------------------------
# 3-state initial temperatures (small, medium, large)
# (order chosen to match d_values and A)
# ------------------------------------------------------------
T0 = Data(
    "Initial Temperatures",
    "T0",
    np.array([
        T_small.value[0],   # small
        T_med.value[0],     # medium
        T_large.value[0],   # large
    ]),
    0.5   # K, broadcast to all three
)

# ------------------------------------------------------------
# Speed of sound (3-state) via RSS: a = sqrt(gamma * R * T0)
# ------------------------------------------------------------
ai = RSS("Speed of Sound", equations[2], [gamma, R, T0])
# ai.var.name -> "a"
# ai.value[0] -> small, ai.value[1] -> medium, ai.value[2] -> large

# ------------------------------------------------------------
# Characteristic time (3-state): tchar = V / (A * a)
# A is [small, medium, large] from d_values
# ------------------------------------------------------------
tchar = RSS("Characteristic Time", equations[3], [V, A, ai])
# tchar.value[0] -> small, [1] -> medium, [2] -> large


#find choke points

P_choke = (1 + 0.5 * (gamma.value - 1)) ** (gamma.value / (gamma.value - 1))

for pressure in Pabs_small.value:
    if pressure / P_amb.value <= P_choke:
        print(f"Small orifice is choked at pressure: {pressure:.2f} Pa")
        choke_index = np.where(Pabs_small.value == pressure)[0][0]
        print(f"Choke occurs at time: {t_small.value[choke_index]:.2f} s")
        break

for pressure in Pabs_med.value:
    if pressure / P_amb.value <= P_choke:
        print(f"Medium orifice is choked at pressure: {pressure:.2f} Pa")
        choke_index = np.where(Pabs_med.value == pressure)[0][0]
        print(f"Choke occurs at time: {t_med.value[choke_index]:.2f} s")
        break

for pressure in Pabs_large.value:
    if pressure / P_amb.value <= P_choke:
        print(f"Large orifice is choked at pressure: {pressure:.2f} Pa")
        choke_index = np.where(Pabs_large.value == pressure)[0][0]
        print(f"Choke occurs at time: {t_large.value[choke_index]:.2f} s")
        break

import os
import numpy as np
import matplotlib.pyplot as plt

# ------------------------------------------------------------
# Build nondimensional P+ and t+ with error bars, per orifice
# ------------------------------------------------------------

P_choke = (1 + 0.5*(gamma.value - 1.0))**(gamma.value/(gamma.value - 1.0))
P_atm   = P_amb.value  # Pa

def nondim_with_errors(tD: Data, PabsD: Data, tchar_scalar: float, label: str):
    """
    tD      : Data object for time
    PabsD   : Data object for absolute pressure (with uncertainties)
    tchar_scalar : characteristic time for that orifice
    label   : 'Small' / 'Medium' / 'Large'

    Returns:
      t_plus, P_plus, sigma_P_plus
    """
    t    = np.asarray(tD.value, dtype=float)
    P    = np.asarray(PabsD.value, dtype=float)
    sigP = np.asarray(PabsD.uncertainty, dtype=float)

    # Choked region: P/P_atm >= P_choke
    ratio      = P / P_atm
    mask_choke = ratio >= P_choke

    if not np.any(mask_choke):
        print(f"[{label}] WARNING: no choked points found.")
        return np.array([]), np.array([]), np.array([])

    t_ch    = t[mask_choke]
    P_ch    = P[mask_choke]
    sigP_ch = sigP[mask_choke]

    # Reference pressure P0 and its uncertainty (first choked point)
    P0    = P_ch[0]
    sigP0 = sigP_ch[0]

    print(f"[{label}] P0 = {P0:.2f} Pa, t_char = {tchar_scalar:.3f} s")

    # Nondimensional time
    t_plus = t_ch / tchar_scalar

    # Nondimensional pressure
    P_plus = P_ch / P0

    # Uncertainty in P_plus via linear propagation:
    # P+ = P / P0  ->  σ_P+ ≈ P+ * sqrt( (σ_P/P)^2 + (σ_P0/P0)^2 )
    eps = 1e-12
    rel_var_P  = (sigP_ch / np.maximum(np.abs(P_ch), eps))**2
    rel_var_P0 = (sigP0   / max(abs(P0), eps))**2
    sigma_P_plus = P_plus * np.sqrt(rel_var_P + rel_var_P0)

    return t_plus, P_plus, sigma_P_plus


# Extract scalar tchar for each orifice (ordering: small, medium, large)
tchar_small = float(tchar.value[0])
tchar_med   = float(tchar.value[1])
tchar_large = float(tchar.value[2])

# (Optional) force evaluation of these uncertainties
tchar.get_error()
slope.get_error()
intercept.get_error()

here = os.path.dirname(os.path.abspath(__file__))

# ------------------------------------------------------------
# Precompute theory constants (same gamma for all orifices)
# ------------------------------------------------------------

gamma_val = float(gamma.value)

# Constant C = ((γ+1)/2)^{-(γ+1)/(2(γ-1))}
C = ((gamma_val + 1.0) / 2.0) ** (-(gamma_val + 1.0) / (2.0 * (gamma_val - 1.0)))

def add_theory_curves(t_plus_data: np.ndarray):
    if t_plus_data.size == 0:
        return

    tmax = float(np.max(t_plus_data))
    tplus_theory = np.linspace(0.0, max(1.2 * tmax, 1.0), 400)

    Pplus_iso = np.exp(-C * tplus_theory)
    Pplus_adi = (
        1.0 + 0.5 * (gamma_val - 1.0) * C * tplus_theory
    ) ** (-2.0 * gamma_val / (gamma_val - 1.0))

    # Plot ADIABATIC on top in RED
    plt.plot(tplus_theory, Pplus_adi, "-",
             color="red", linewidth=2, zorder=5,
             label="Choked adiabatic theory")

    # Plot ISOTHERMAL slightly below it
    plt.plot(tplus_theory, Pplus_iso, "--",
             linewidth=2, zorder=4,
             label="Choked isentropic theory")



# ------------------------------------------------------------
# Small orifice: data + theory
# ------------------------------------------------------------
ts, Ps, sigPs = nondim_with_errors(t_small, Pabs_small, tchar_small, "Small")
plt.figure(figsize=(7, 5))

# 1️⃣ Plot data FIRST (under everything else)
if ts.size > 0:
    plt.errorbar(ts, Ps, yerr=sigPs, fmt='o', capsize=3,
                 label="Small Orifice data",
                 zorder=1)   # << LOWER than theory

# 2️⃣ Then theory curves
add_theory_curves(ts)



plt.xlabel(r"$t^+ = t / t_{\rm char}$")
plt.ylabel(r"$P^+ = P / P_0$")
plt.title("Non-Dimensional Choked Tank Discharge – Small Orifice")
plt.grid(True, linestyle="--", alpha=0.5)
plt.legend()
plt.tight_layout()
plt.savefig(os.path.join(here, "lab6_nondim_small_with_theory.png"), dpi=300)
plt.show()

# ------------------------------------------------------------
# Medium orifice: data + theory
# ------------------------------------------------------------
# MEDIUM
tm, Pm, sigPm = nondim_with_errors(t_med, Pabs_med, tchar_med, "Medium")
plt.figure(figsize=(7, 5))
if tm.size > 0:
    plt.errorbar(tm, Pm, yerr=sigPm, fmt='s', capsize=3,
                 label="Medium Orifice data",
                 zorder=1)
add_theory_curves(tm)



plt.xlabel(r"$t^+ = t / t_{\rm char}$")
plt.ylabel(r"$P^+ = P / P_0$")
plt.title("Non-Dimensional Choked Tank Discharge – Medium Orifice")
plt.grid(True, linestyle="--", alpha=0.5)
plt.legend()
plt.tight_layout()
plt.savefig(os.path.join(here, "lab6_nondim_medium_with_theory.png"), dpi=300)
plt.show()

# ------------------------------------------------------------
# Large orifice: data + theory
# ------------------------------------------------------------
# LARGE
tl, Pl, sigPl = nondim_with_errors(t_large, Pabs_large, tchar_large, "Large")
plt.figure(figsize=(7, 5))
if tl.size > 0:
    plt.errorbar(tl, Pl, yerr=sigPl, fmt='^', capsize=3,
                 label="Large Orifice data",
                 zorder=1)
add_theory_curves(tl)



plt.xlabel(r"$t^+ = t / t_{\rm char}$")
plt.ylabel(r"$P^+ = P / P_0$")
plt.title("Non-Dimensional Choked Tank Discharge – Large Orifice")
plt.grid(True, linestyle="--", alpha=0.5)
plt.legend()
plt.tight_layout()
plt.savefig(os.path.join(here, "lab6_nondim_large_with_theory.png"), dpi=300)
plt.show()


# ------------------------------------------------------------
# Temperature vs nondimensional time (all three orifices)
# ------------------------------------------------------------

tplus_T_small = np.asarray(t_small.value, dtype=float) / tchar_small
tplus_T_med   = np.asarray(t_med.value,   dtype=float) / tchar_med
tplus_T_large = np.asarray(t_large.value, dtype=float) / tchar_large

plt.figure(figsize=(8, 6))

plt.plot(tplus_T_small, T_small.value, 'o-', label="Small Orifice")
plt.plot(tplus_T_med,   T_med.value,   's-', label="Medium Orifice")
plt.plot(tplus_T_large, T_large.value, '^-', label="Large Orifice")

plt.axhline(T_amb.value, linestyle="--", alpha=0.6,
            label=f"Ambient $T_{{amb}} \\approx {T_amb.value:.1f}\\,\\mathrm{{K}}$")

plt.xlabel(r"$t^+ = t / t_{\rm char}$")
plt.ylabel(r"Temperature $T$ [K]")
plt.title("Tank Temperature vs Nondimensional Time – All Orifices")
plt.grid(True, linestyle="--", alpha=0.5)
plt.legend()

# *** UPDATED AXIS RANGE ***
plt.ylim(273, 320)

plt.tight_layout()
plt.savefig(os.path.join(here, "lab6_temperature_vs_tchar.png"), dpi=300)
plt.show()


